#!/usr/bin/env python3
'''
from openpyxl import load_workbook
load_wb = load_workbook("./FOCUS_MODEL_DEFINE.xlsx", data_only=True)
load_ws = load_wb['Sheet1']


print(load_ws['A1'].value)

print(load_ws.cell(1, 2).value)
'''
import openpyxl
import os

OEM_PATH = "./define/model_spec/"
OEM_WHOLE_PATH_NAME = ""


def getOemName(row):
	column = 1
	key = 0

	while key is not None:
		key = sheet.cell(row, column).value
		print("key = ", key)
		if key == "OEM":
			sValue = sheet.cell(row + 1, column).value
			print("sValue = ", sValue)
			return str(sValue)
		column += 1


def makeDirectory():
	try:
		print("RUN getOemName1")
		OEM_NAME = getOemName(1)
		print("OEM_NAME = ", OEM_NAME)
		print("RUN getOemName2")
		global OEM_WHOLE_PATH_NAME
		OEM_WHOLE_PATH_NAME = OEM_PATH + OEM_NAME
		print("OEM_WHOLE_PATH_NAME = ", OEM_WHOLE_PATH_NAME)
		if not(os.path.isdir(OEM_WHOLE_PATH_NAME)):
			os.makedirs(OEM_WHOLE_PATH_NAME)
	except OSError as e:
		if e.error != errno.EEXIST:
			print("Failed to create directory")
			raise


def gen_file(row):
	column = 1
	key = 0
	while key is not None:
		key = sheet.cell(row, column).value
		print("key = ", key)
		if key == "MODEL":
			print("OEM_WHOLE_PATH_NAME = ", OEM_WHOLE_PATH_NAME)
			sValue = OEM_WHOLE_PATH_NAME + str("/") + sheet.cell(row + 1, column).value
			print("sValue = ", sValue)
			return open(sValue, 'w')
		column += 1
	return None


filename = "./FOCUS_DEFINE_MODEL.xlsx"
book = openpyxl.load_workbook(filename)
sheet = book['Sheet2']

sKey = " "
section = 1
index = 1

print("RUN makeDirectory1")
makeDirectory()
print("RUN makeDirectory2")
print("RUN gen_file1")
ini_file = gen_file(index)
print("RUN gen_file2")

while sKey is not None:
	sKey = sheet.cell(section, index).value
	value = sheet.cell(section + 1, index).value
	if sKey is not None and ini_file is not None:
		data = str(sKey) + '=' + str(value) + '\n'
		if(ini_file.writable()):
			ini_file.write(data)
	print("sheet.cell([row = ", section, ", column = ", index, ").value = ", sKey, ":", value)
	index += 1
